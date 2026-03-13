import io
import logging
from datetime import date, datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)

from app.models import BorrowTransaction

logger = logging.getLogger(__name__)


def export_to_excel(transactions: List[BorrowTransaction]) -> bytes:
    """
    Export transactions to an Excel workbook.
    Returns the workbook as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Borrowing Report"

    # ── Styles ────────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Title Row ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"IT Equipment Borrowing Report - Generated {date.today().isoformat()}"
    title_cell.font = Font(name="Calibri", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # ── Headers ───────────────────────────────────────────────────────────
    headers = [
        "Transaction ID", "NIK", "Employee Name", "Department",
        "Items Borrowed", "Borrow Date", "Return Date",
        "Status", "Notes"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── Data Rows ─────────────────────────────────────────────────────────
    status_fills = {
        "BORROWED": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "RETURNED": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "OVERDUE": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    }

    for row_idx, txn in enumerate(transactions, 4):
        # Build items string
        items_str = ", ".join(
            f"{d.item.item_name} (x{d.quantity})"
            for d in txn.details if d.item
        )

        employee_name = txn.employee.name if txn.employee else "N/A"
        department = txn.employee.department if txn.employee else "N/A"
        status_value = txn.status.value if txn.status else "N/A"

        row_data = [
            txn.id,
            txn.nik,
            employee_name,
            department,
            items_str,
            txn.borrow_date.isoformat() if txn.borrow_date else "",
            txn.return_date.isoformat() if txn.return_date else "",
            status_value,
            txn.notes or "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Color-code status column
            if col_idx == 8 and status_value in status_fills:
                cell.fill = status_fills[status_value]

    # ── Column Widths ─────────────────────────────────────────────────────
    column_widths = [16, 14, 22, 18, 35, 14, 14, 12, 30]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[
            ws.cell(row=1, column=idx).column_letter
        ].width = width

    # ── Summary Row ───────────────────────────────────────────────────────
    summary_row = len(transactions) + 5
    ws.cell(row=summary_row, column=1, value=f"Total Transactions: {len(transactions)}").font = Font(bold=True)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_to_pdf(transactions: List[BorrowTransaction]) -> bytes:
    """
    Export transactions to a PDF document.
    Returns the PDF as bytes.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    elements = []
    styles = getSampleStyleSheet()

    # ── Title ─────────────────────────────────────────────────────────────
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=12,
        alignment=1,  # Center
    )
    elements.append(Paragraph("IT Equipment Borrowing Report", title_style))

    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=10,
        spaceAfter=20,
        alignment=1,
        textColor=colors.grey,
    )
    elements.append(Paragraph(
        f"Generated on {date.today().isoformat()}", subtitle_style
    ))
    elements.append(Spacer(1, 10))

    # ── Table Data ────────────────────────────────────────────────────────
    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
    )

    header_data = [
        "ID", "NIK", "Employee", "Dept",
        "Items", "Borrow Date", "Return Date", "Status", "Notes"
    ]

    table_data = [header_data]

    for txn in transactions:
        items_str = ", ".join(
            f"{d.item.item_name} (x{d.quantity})"
            for d in txn.details if d.item
        )

        employee_name = txn.employee.name if txn.employee else "N/A"
        department = txn.employee.department if txn.employee else "N/A"
        status_value = txn.status.value if txn.status else "N/A"

        row = [
            str(txn.id),
            txn.nik,
            Paragraph(employee_name, cell_style),
            department,
            Paragraph(items_str, cell_style),
            txn.borrow_date.isoformat() if txn.borrow_date else "",
            txn.return_date.isoformat() if txn.return_date else "",
            status_value,
            Paragraph(txn.notes or "-", cell_style),
        ]
        table_data.append(row)

    # If no data, add empty row
    if len(table_data) == 1:
        table_data.append(["No transactions found"] + [""] * 8)

    # ── Table Styling ─────────────────────────────────────────────────────
    col_widths = [30, 55, 75, 55, 120, 60, 60, 50, 100]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B579A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

        # Body
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),  # ID centered

        # Grid
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#2B579A")),

        # Alternating row colors
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),

        # Padding
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    # Color-code status cells
    status_colors = {
        "BORROWED": colors.HexColor("#FFF3CD"),
        "RETURNED": colors.HexColor("#D4EDDA"),
        "OVERDUE": colors.HexColor("#F8D7DA"),
    }
    for row_idx, txn in enumerate(transactions, 1):
        status_val = txn.status.value if txn.status else ""
        if status_val in status_colors:
            table.setStyle(TableStyle([
                ("BACKGROUND", (7, row_idx), (7, row_idx), status_colors[status_val]),
            ]))

    elements.append(table)

    # ── Footer / Summary ──────────────────────────────────────────────────
    elements.append(Spacer(1, 20))
    summary_style = ParagraphStyle(
        "Summary",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.grey,
    )
    elements.append(Paragraph(
        f"Total Transactions: {len(transactions)}", summary_style
    ))

    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
